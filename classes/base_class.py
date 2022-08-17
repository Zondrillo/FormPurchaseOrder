import os
import tempfile as tf
import xlsxwriter as xl
from openpyxl import load_workbook
from pandas import DataFrame

from configs import config, texts
from utilities.helpers import get_supply_months


class BaseClass:

    def __init__(self, pivot_table: DataFrame):
        self.file_path = tf.mktemp(suffix='.xlsx',
                                   dir='')  # создаём временный файл для записи сводной таблицы одного завода
        pivot_table.to_excel(self.file_path, merge_cells=False)  # преобразовываем сводную таблицу в формат excel
        self.temp_wb = load_workbook(filename=self.file_path,
                                     data_only=True)  # получаем данные из excel-файла со сводной таблицей
        self.temp_ws = self.temp_wb.active  # выбираем единственный временный лист в excel-файле
        self.factory_id = self.temp_ws['B2'].value  # получаем id завода, с которым работаем в данный момент
        self.budget_name = self.temp_ws['A2'].value  # получаем раздел ГКПЗ с которым работаем в данный момент
        self.final_wb = xl.Workbook()  # создаём конечный excel-файл, в который будем записывать данные
        self.final_ws = self.final_wb.add_worksheet()  # добавляем лист, в который будем записывать данные
        self.final_ws.set_landscape()  # альбомная ориентация
        self.final_ws.set_paper(9)  # формат А4
        self.final_ws.fit_to_pages(1, 0)  # вписать все столбцы на одну страницу
        self.final_ws.set_zoom(60)  # установить масштаб 60%

    def make_head(self) -> None:
        """Формирует шапку ТЗ"""
        format_head = self.final_wb.add_format(config.format_head)
        merge_format1 = self.final_wb.add_format(config.merge_format1)
        merge_format2 = self.final_wb.add_format(config.merge_format2)
        merge_format3 = self.final_wb.add_format(config.merge_format3)
        rotate = self.final_wb.add_format(config.rotate_format)
        self.final_ws.set_column('A:A', 6)
        self.final_ws.set_column('B:C', 13.5)
        self.final_ws.set_column('D:D', 43)
        self.final_ws.set_column('E:E', 54)
        self.final_ws.set_column('F:F', 9.5)
        self.final_ws.set_column('G:G', 18)
        self.final_ws.set_column('H:T', 15)
        self.final_ws.write('U1', 'Приложение № 2 к Приказу НФ "ПАО "Т Плюс"', format_head)
        self.final_ws.write('U2', '№___________________________________________ от ____________________________',
                            format_head)
        self.final_ws.merge_range('A4:U4', f'Техническое задание на поставку {config.lot_name}', merge_format2)
        self.final_ws.merge_range('A5:C5', 'Таблица 1', merge_format3)
        col_head = 0
        for element in config.tech_task_head:
            self.final_ws.merge_range(5, col_head, 6, col_head, element, merge_format1)
            col_head += 1
        months = get_supply_months()
        col_month = 7
        for month in months:
            self.final_ws.write_datetime(6, col_month, month, rotate)
            col_month += 1
        self.final_ws.merge_range('H6:T6', 'Срок поставки', merge_format1)
        self.final_ws.merge_range('U6:U7', 'Грузополучатель', merge_format1)

    def make_tail(self, row_num: int) -> None:
        """Вставляет таблицу № 2 в ТЗ"""
        merge_format1 = self.final_wb.add_format(config.merge_format3)
        merge_format2 = self.final_wb.add_format(config.merge_format1)
        self.merge_format3 = self.final_wb.add_format(config.merge_format4)
        self.final_ws.merge_range(f'A{row_num}:C{row_num}', 'Таблица 2', merge_format1)
        self.final_ws.write_string(f'A{row_num + 1}', '№ п/п', merge_format2)
        self.final_ws.merge_range(f'B{row_num + 1}:D{row_num + 1}', 'Показатель', merge_format2)
        self.final_ws.merge_range(f'E{row_num + 1}:U{row_num + 1}', 'Описание', merge_format2)
        self.final_ws.merge_range(f'A{row_num + 2}:A{row_num + 6}', 1, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 2}:D{row_num + 6}', texts.supply_conditions_title, self.merge_format3)
        self.final_ws.merge_range(f'E{row_num + 2}:U{row_num + 2}', texts.supply_conditions_desc1, self.merge_format3)
        self.final_ws.set_row(row_num + 1, 60)
        self.final_ws.merge_range(f'E{row_num + 4}:U{row_num + 4}', texts.supply_conditions_desc3, self.merge_format3)
        self.final_ws.merge_range(f'E{row_num + 5}:U{row_num + 5}', texts.supply_conditions_desc4, self.merge_format3)
        self.final_ws.set_row(row_num + 4, 148.20)
        self.final_ws.merge_range(f'E{row_num + 6}:U{row_num + 6}', texts.supply_conditions_desc5, self.merge_format3)
        self.final_ws.merge_range(f'A{row_num + 7}:A{row_num + 10}', 2, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 7}:D{row_num + 10}', texts.quality_requirements_title,
                                  self.merge_format3)
        self.final_ws.merge_range(f'E{row_num + 7}:U{row_num + 7}', texts.quality_requirements_desc1,
                                  self.merge_format3)
        self.final_ws.merge_range(f'E{row_num + 8}:U{row_num + 8}', texts.quality_requirements_desc2,
                                  self.merge_format3)
        self.final_ws.set_row(row_num + 7, 40)
        self.final_ws.merge_range(f'E{row_num + 9}:U{row_num + 9}', texts.quality_requirements_desc3,
                                  self.merge_format3)
        self.final_ws.merge_range(f'E{row_num + 10}:U{row_num + 10}', 'Иное: нет', self.merge_format3)
        self.final_ws.write_number(f'A{row_num + 11}', 3, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 11}:D{row_num + 11}', texts.confirmation_of_compliance_title,
                                  self.merge_format3)
        self.final_ws.merge_range(f'E{row_num + 11}:U{row_num + 11}', texts.confirmation_of_compliance_desc1,
                                  self.merge_format3)
        self.final_ws.set_row(row_num + 10, 81)
        self.final_ws.merge_range(f'A{row_num + 12}:A{row_num + 14}', 4, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 12}:D{row_num + 14}', texts.safety_requirements_title,
                                  self.merge_format3)
        self.final_ws.merge_range(f'E{row_num + 12}:U{row_num + 12}', texts.safety_requirements_desc1,
                                  self.merge_format3)
        self.final_ws.merge_range(f'E{row_num + 13}:U{row_num + 13}', texts.safety_requirements_desc2,
                                  self.merge_format3)
        self.final_ws.merge_range(f'E{row_num + 14}:U{row_num + 14}', 'Иное: нет', self.merge_format3)
        self.final_ws.merge_range(f'A{row_num + 15}:A{row_num + 18}', 5, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 15}:C{row_num + 18}', 'Иные требования', self.merge_format3)
        self.final_ws.write_string(f'D{row_num + 15}', 'Эквивалент', self.merge_format3)
        self.final_ws.write_string(f'D{row_num + 16}', 'Толеранс (+/-), %', self.merge_format3)
        self.final_ws.write_string(f'D{row_num + 17}', 'Срок службы (расчетный ресурс)', self.merge_format3)
        self.final_ws.write_string(f'D{row_num + 18}', 'Другое', self.merge_format3)
        self.final_ws.merge_range(f'E{row_num + 15}:U{row_num + 15}', texts.equivalent_desc, self.merge_format3)
        self.final_ws.set_row(row_num + 14, 42.6)
        self.final_ws.merge_range(f'E{row_num + 16}:U{row_num + 16}', 'Нет', self.merge_format3)
        self.final_ws.merge_range(f'E{row_num + 17}:U{row_num + 17}', None, self.merge_format3)
        self.final_ws.merge_range(f'E{row_num + 18}:U{row_num + 18}', 'Нет', self.merge_format3)

    def close_and_clear(self):
        self.temp_wb.close()
        self.final_wb.close()
        os.remove(self.file_path)
