import os
import tempfile as tf
import xlsxwriter as xl
from openpyxl import load_workbook
from pandas import DataFrame

from configs import config, texts
from utilities.helpers import get_supply_months, get_words_in_genitive_case


class FormTechTaskComm:

    def __init__(self, some_table: DataFrame, some_list: list):
        self.file_path = tf.mktemp(suffix='.xlsx', dir='')  # создаём временный файл для записи сводной таблицы одного завода
        some_table.to_excel(self.file_path, merge_cells=False)  # преобразовываем сводную таблицу в формат excel
        self.some_list = some_list  # номера строк для подсчёта итога по станции/сетям
        self.temp_wb = load_workbook(filename=self.file_path, data_only=True)  # получаем данные из excel-файла со сводной таблицей
        self.temp_ws = self.temp_wb.active  # выбираем единственный временный лист в excel-файле
        budget_name = self.temp_ws['A2'].value  # запоминаем раздел ГКПЗ с которым работаем в данный момент
        self.final_wb = xl.Workbook(f'ТЗ_{budget_name}_{config.lot_name}_Общее.xlsx')  # создаём конечный excel-файл, в который будем записывать данные
        self.final_ws = self.final_wb.add_worksheet(f'{budget_name}')  # добавляем лист, в который будем записывать данные
        self.final_ws.set_landscape()  # альбомная ориентация
        self.final_ws.set_paper(9)  # формат А4
        self.final_ws.fit_to_pages(1, 0)  # вписать все столбцы на одну страницу
        self.final_ws.set_zoom(60)  # установить масштаб 60%

    def big_table(self) -> list:
        """Получает все необходимые данные из сводной таблицы"""
        lst = []
        i = 1
        j = 0
        for row in self.temp_ws.iter_rows(min_row=2, max_row=self.temp_ws.max_row, min_col=2, max_col=19):
            lst.append([i] + [cell.value for cell in row])
            if i < self.some_list[j]:
                i += 1
            else:
                i = 1
                j += 1
        [element.insert(5, None) for element in lst]  # вставляет пустой столбец для технических требований
        return lst

    def make_middle(self, lst_data: list) -> tuple:
        """Добавляет данные в таблицу 1 ТЗ"""
        r_num = 8
        format_pivot_table = self.final_wb.add_format(config.format_pivot_table)
        quantity_format = self.final_wb.add_format(config.quantity_format)
        format_total_text = self.final_wb.add_format(config.format_total_text)
        format_total_num = self.final_wb.add_format(config.format_total_num)
        self.final_ws.set_column('U:U', 46)
        k = 0
        factories_set = set()
        for row in lst_data:
            factory_id = row[1]
            factories_set.add(factory_id)
            self.final_ws.write_number(f'A{r_num}', row[0], format_pivot_table)
            self.final_ws.write_row(f'B{r_num}', row[2:7], format_pivot_table)
            self.final_ws.write_formula(r_num - 1, 6, f'=SUM(H{r_num}:T{r_num})', quantity_format)
            self.final_ws.write_row(f'H{r_num}', row[7:], quantity_format)
            self.final_ws.write_string(f'U{r_num}', texts.addresses[f'{factory_id}'], format_pivot_table)
            if row[0] == self.some_list[k]:
                self.final_ws.merge_range(r_num, 0, r_num, 5, texts.totals[f'{factory_id}'], format_total_text)
                for cell in config.cells:
                    self.final_ws.write_formula(f'{cell}{r_num + 1}',
                                                f'=SUM({cell}{r_num - self.some_list[k] + 1}:{cell}{r_num})',
                                                format_total_num)
                r_num += 1
                k += 1
                self.final_ws.write(f'U{r_num}', None, format_total_text)
            r_num += 1
        self.final_ws.merge_range(r_num - 1, 0, r_num - 1, 5, 'Общий итог', format_total_text)
        for cell in config.cells:
            next_total_row_number = 8
            formula_constructor = '='
            for index, total_row_number in enumerate(self.some_list):
                next_total_row_number += total_row_number
                formula_constructor += f'{cell}{next_total_row_number + index}+'
            self.final_ws.write_formula(f'{cell}{r_num}', formula_constructor[:-1], format_total_num)
        self.final_ws.write(f'U{r_num}', None, format_total_text)
        return r_num, factories_set

    def make_head(self) -> None:
        """Формирует шапку ТЗ"""
        format_head = self.final_wb.add_format(config.format_head)
        merge_format1 = self.final_wb.add_format(config.merge_format1)
        merge_format2 = self.final_wb.add_format(config.merge_format2)
        merge_format3 = self.final_wb.add_format(config.merge_format3)
        rotate = self.final_wb.add_format(config.rotate_format)
        self.final_ws.set_column('A:A', 6)
        self.final_ws.set_column('B:C', 13.5)
        self.final_ws.set_column('D:D', 43)
        self.final_ws.set_column('E:E', 54)
        self.final_ws.set_column('F:F', 9.5)
        self.final_ws.set_column('G:G', 18)
        self.final_ws.set_column('H:T', 15)
        self.final_ws.write('U1', 'Приложение № 2 к Приказу НФ "ПАО "Т Плюс"', format_head)
        self.final_ws.write('U2', '№___________________________________________ от ____________________________',
                            format_head)
        self.final_ws.merge_range('A4:U4',
                                  f'Техническое задание на поставку {get_words_in_genitive_case(config.lot_name)}',
                                  merge_format2)
        self.final_ws.merge_range('A5:C5', 'Таблица 1', merge_format3)
        col_head = 0
        for element in config.head:
            self.final_ws.merge_range(5, col_head, 6, col_head, element, merge_format1)
            col_head += 1
        months = get_supply_months(config.start_month)
        col_month = 7
        for month in months:
            self.final_ws.write_datetime(6, col_month, month, rotate)
            col_month += 1
        self.final_ws.merge_range('H6:T6', 'Срок поставки', merge_format1)
        self.final_ws.merge_range('U6:U7', 'Грузополучатель', merge_format1)

    def make_tail(self, row_num: int, factory_id_set: set) -> None:
        """Вставляет таблицу № 2 в ТЗ"""
        merge_format1 = self.final_wb.add_format(config.merge_format3)
        merge_format2 = self.final_wb.add_format(config.merge_format1)
        merge_format3 = self.final_wb.add_format(config.merge_format4)
        self.final_ws.merge_range(f'A{row_num}:C{row_num}', 'Таблица 2', merge_format1)
        self.final_ws.write_string(f'A{row_num + 1}', '№ п/п', merge_format2)
        self.final_ws.merge_range(f'B{row_num + 1}:D{row_num + 1}', 'Показатель', merge_format2)
        self.final_ws.merge_range(f'E{row_num + 1}:U{row_num + 1}', 'Описание', merge_format2)
        self.final_ws.merge_range(f'A{row_num + 2}:A{row_num + 6}', 1, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 2}:D{row_num + 6}', texts.supply_conditions_title, merge_format3)
        self.final_ws.merge_range(f'E{row_num + 2}:U{row_num + 2}', texts.supply_conditions_desc1, merge_format3)
        self.final_ws.set_row(row_num + 1, 60)
        adresses_string = ''
        for factory_id in factory_id_set:  # подготавливаем список грузополучателей
            adresses_string += f'{texts.addresses[factory_id]}\n'
        self.final_ws.merge_range(f'E{row_num + 3}:U{row_num + 3}',
                                  texts.supply_conditions_desc2 + adresses_string.strip(), merge_format3)
        self.final_ws.set_row(row_num + 2, config.addresses_row_height[len(factory_id_set) - 1])
        self.final_ws.merge_range(f'E{row_num + 4}:U{row_num + 4}', texts.supply_conditions_desc3, merge_format3)
        self.final_ws.merge_range(f'E{row_num + 5}:U{row_num + 5}', texts.supply_conditions_desc4, merge_format3)
        self.final_ws.set_row(row_num + 4, 148.20)
        self.final_ws.merge_range(f'E{row_num + 6}:U{row_num + 6}', texts.supply_conditions_desc5, merge_format3)
        self.final_ws.merge_range(f'A{row_num + 7}:A{row_num + 10}', 2, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 7}:D{row_num + 10}', texts.quality_requirements_title, merge_format3)
        self.final_ws.merge_range(f'E{row_num + 7}:U{row_num + 7}', texts.quality_requirements_desc1, merge_format3)
        self.final_ws.merge_range(f'E{row_num + 8}:U{row_num + 8}', texts.quality_requirements_desc2, merge_format3)
        self.final_ws.set_row(row_num + 7, 40)
        self.final_ws.merge_range(f'E{row_num + 9}:U{row_num + 9}', texts.quality_requirements_desc3, merge_format3)
        self.final_ws.merge_range(f'E{row_num + 10}:U{row_num + 10}', 'Иное: нет', merge_format3)
        self.final_ws.write_number(f'A{row_num + 11}', 3, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 11}:D{row_num + 11}', texts.confirmation_of_compliance_title,
                                  merge_format3)
        self.final_ws.merge_range(f'E{row_num + 11}:U{row_num + 11}', texts.confirmation_of_compliance_desc1,
                                  merge_format3)
        self.final_ws.set_row(row_num + 10, 81)
        self.final_ws.merge_range(f'A{row_num + 12}:A{row_num + 14}', 4, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 12}:D{row_num + 14}', texts.safety_requirements_title, merge_format3)
        self.final_ws.merge_range(f'E{row_num + 12}:U{row_num + 12}', texts.safety_requirements_desc1, merge_format3)
        self.final_ws.merge_range(f'E{row_num + 13}:U{row_num + 13}', texts.safety_requirements_desc2, merge_format3)
        self.final_ws.merge_range(f'E{row_num + 14}:U{row_num + 14}', 'Иное: нет', merge_format3)
        self.final_ws.merge_range(f'A{row_num + 15}:A{row_num + 18}', 5, merge_format2)
        self.final_ws.merge_range(f'B{row_num + 15}:C{row_num + 18}', 'Иные требования', merge_format3)
        self.final_ws.write_string(f'D{row_num + 15}', 'Эквивалент', merge_format3)
        self.final_ws.write_string(f'D{row_num + 16}', 'Толеранс (+/-), %', merge_format3)
        self.final_ws.write_string(f'D{row_num + 17}', 'Срок службы (расчетный ресурс)', merge_format3)
        self.final_ws.write_string(f'D{row_num + 18}', 'Другое', merge_format3)
        self.final_ws.merge_range(f'E{row_num + 15}:U{row_num + 15}', texts.equivalent_desc, merge_format3)
        self.final_ws.set_row(row_num + 14, 42.6)
        self.final_ws.merge_range(f'E{row_num + 16}:U{row_num + 16}', 'Нет', merge_format3)
        self.final_ws.merge_range(f'E{row_num + 17}:U{row_num + 17}', None, merge_format3)
        self.final_ws.merge_range(f'E{row_num + 18}:U{row_num + 18}', 'Нет', merge_format3)

    def form(self) -> None:
        """Формирует ТЗ"""
        self.make_head()
        curr_row_num = self.make_middle(self.big_table())
        self.make_tail(curr_row_num[0] + 2, curr_row_num[1])
        self.temp_wb.close()
        self.final_wb.close()
        os.remove(self.file_path)
